from __future__ import annotations
import io
from typing import Any, Callable, Dict, Optional
import pandas


def apply_excel_style(
    filename_or_writer: Any,
    f_highlights: Optional[  # type: ignore[name-defined]
        Dict[
            str,
            Callable[
                [Any], CubeViewDef.HighLightKind  # pyrefly: ignore[unknown-name]]
            ],
        ]
    ] = None,
    time_mask_view: Optional[Dict[str, pandas.DataFrame]] = None,
    verbose: int = 0,
):
    """
    Applies styles on all sheets in a file unless the sheet is too big.

    :param filename_or_writer: filename, modified inplace
    :param f_highlight: color function to apply, one per sheet
    :param time_mask_view: if specified, it contains dataframe with the same shape
        and values in {-1, 0, +1} which indicates if a value is unexpectedly lower (-1)
        or higher (+1), it changes the color of the background then.
    :param verbose: progress loop
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill
    from .cube_helper import CubeViewDef

    if isinstance(filename_or_writer, str):
        workbook = load_workbook(filename_or_writer)
        save = True
    else:
        workbook = filename_or_writer.book
        save = False

    mask_low = PatternFill(fgColor="AAAAF0", fill_type="solid")
    mask_high = PatternFill(fgColor="F0AAAA", fill_type="solid")

    left = Alignment(horizontal="left")
    left_shrink = Alignment(horizontal="left", shrink_to_fit=True)
    right = Alignment(horizontal="right")
    font_colors = {
        CubeViewDef.HighLightKind.GREEN: Font(color="00AA00"),
        CubeViewDef.HighLightKind.RED: Font(color="FF0000"),
    }
    if verbose:  # pragma: no cover
        from tqdm import tqdm

        sheet_names = tqdm(list(workbook.sheetnames))
    else:
        sheet_names = workbook.sheetnames
    for name in sheet_names:
        if time_mask_view and name in time_mask_view:
            mask = time_mask_view[name]
            with pandas.ExcelWriter(io.BytesIO(), engine="openpyxl") as mask_writer:
                mask.to_excel(mask_writer, sheet_name=name)
                sheet_mask = mask_writer.sheets[name]
        else:
            sheet_mask = None

        f_highlight = f_highlights.get(name, None) if f_highlights else None
        sheet = workbook[name]
        n_rows = sheet.max_row
        n_cols = sheet.max_column
        if n_rows * n_cols > 2**16 or n_rows > 2**13:
            # Too big.
            continue
        co: Dict[int, int] = {}
        sizes: Dict[int, int] = {}
        cols = set()
        for i in range(1, n_rows + 1):
            for j, cell in enumerate(sheet[i]):
                if j > n_cols:
                    break
                cols.add(cell.column)
                if isinstance(cell.value, float):
                    co[j] = co.get(j, 0) + 1
                elif isinstance(cell.value, str):
                    sizes[cell.column] = max(sizes.get(cell.column, 0), len(cell.value))

        for k, v in sizes.items():
            c = get_column_letter(k)
            sheet.column_dimensions[c].width = min(max(8, v), 30)
        for k in cols:
            if k not in sizes:
                c = get_column_letter(k)
                sheet.column_dimensions[c].width = 15

        for i in range(1, n_rows + 1):
            for j, cell in enumerate(sheet[i]):
                if j > n_cols:
                    break
                if isinstance(cell.value, pandas.Timestamp):
                    cell.alignment = right
                    dt = cell.value.to_pydatetime()
                    cell.value = dt
                    cell.number_format = (
                        "YYYY-MM-DD"
                        if (
                            dt.hour == 0
                            and dt.minute == 0
                            and dt.second == 0
                            and dt.microsecond == 0
                        )
                        else "YYYY-MM-DD 00:00:00"
                    )
                elif isinstance(cell.value, (float, int)):
                    cell.alignment = right
                    x = abs(cell.value)
                    if int(x) == x:
                        cell.number_format = "0"
                    elif x > 5000:
                        cell.number_format = "# ##0"
                    elif x >= 500:
                        cell.number_format = "0.0"
                    elif x >= 50:
                        cell.number_format = "0.00"
                    elif x >= 5:
                        cell.number_format = "0.000"
                    elif x > 0.5:
                        cell.number_format = "0.0000"
                    elif x > 0.005:
                        cell.number_format = "0.00000"
                    else:
                        cell.number_format = "0.000E+00"
                    if f_highlight:
                        h = f_highlight(cell.value)
                        if h in font_colors:
                            cell.font = font_colors[h]
                elif isinstance(cell.value, str) and len(cell.value) > 70:
                    cell.alignment = left_shrink
                else:
                    cell.alignment = left
                    if f_highlight:
                        h = f_highlight(cell.value)
                        if h in font_colors:
                            cell.font = font_colors[h]

        if sheet_mask is not None:
            for i in range(1, n_rows + 1):
                for j, (cell, cell_mask) in enumerate(zip(sheet[i], sheet_mask[i])):
                    if j > n_cols:
                        break
                    if cell_mask.value not in (1, -1):
                        continue
                    cell.fill = mask_low if cell_mask.value < 0 else mask_high

    if save:
        workbook.save(filename_or_writer)
