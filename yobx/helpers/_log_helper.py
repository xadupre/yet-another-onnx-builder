import datetime
import glob
import io
import os
import zipfile
from typing import Any, Callable, Dict, Iterator, List, Optional, Sequence, Tuple, Union
import numpy as np
import pandas

BUCKET_SCALES_VALUES = np.array(
    [-np.inf, -20, -10, -5, -2, 0, 2, 5, 10, 20, 100, 200, 300, 400, np.inf], dtype=float
)


BUCKET_SCALES = BUCKET_SCALES_VALUES / 100 + 1


def mann_kendall(series: Sequence[float], threshold: float = 0.5):
    """
    Computes the test of Mann-Kendall.

    :param series: series
    :param threshold: 1.96 is the usual value, 0.5 means a short timeseries
        ``(0, 1, 2, 3, 4)`` has a significant trend
    :return: trend (-1, 0, +1), test value

    .. math::

        S =\\sum_{i=1}^{n}\\sum_{j=i+1}^{n} sign(x_j - x_i)

    where the function *sign* is:

    .. math::

        sign(x) = \\left\\{ \\begin{array}{l} -1 if x < 0 \\\\ 0 if x = 0 \\\\ +1 otherwise
        \\end{array} \\right.

    And:

    .. math::

        Var(S)= \\frac{n(n-1)(2n+5) - \\sum_t t(t-1)(2t+5)}{18}
    """
    aseries = np.asarray(series)
    stat = 0
    n = len(aseries)
    var = n * (n - 1) * (2 * n + 5) / 18
    for i in range(n - 1):
        stat += np.sign(aseries[i + 1 :] - aseries[i]).sum()
    var = var**0.5
    test = (stat + (1 if stat < 0 else (0 if stat == 0 else -1))) / var
    trend = np.sign(test) if np.abs(test) > threshold else 0
    return trend, test


def breaking_last_point(series: Sequence[float], threshold: float = 1.2):
    """
    Assuming a timeseries is constant, we check the last value
    is not an outlier.

    :param series: series
    :param threshold: number of standard deviations beyond the mean required
        for the last point to be considered a significant change (default 1.2)
    :return: significant change (-1, 0, +1), test value
    """
    signal = np.asarray(series)
    if not np.issubdtype(signal.dtype, np.number):
        return 0, np.nan
    assert len(signal.shape) == 1, f"Unexpected signal shape={signal.shape}, signal={signal}"
    if signal.shape[0] <= 2:
        return 0, 0

    has_value = ~(np.isnan(signal).all()) and ~(np.isinf(signal).all())
    if np.isnan(signal[-1]) or np.isinf(signal[-1]):
        return (-1, np.inf) if has_value else (0, 0)

    try:
        m = np.mean(signal[:-1])
    except (TypeError, ValueError):
        # Not a numerical type
        return 0, np.nan

    if np.isnan(m) or np.isinf(m):
        return (1, np.inf) if np.isinf(signal[-2]) or np.isnan(signal[-2]) else (0, 0)
    v = np.std(signal[:-1])
    if v == 0:
        test = signal[-1] - m
        assert not np.isnan(
            test
        ), f"Unexpected test value, test={test}, signal={signal}, m={m}, v={v}"
        trend = np.sign(test)
        return trend, trend
    test = (signal[-1] - m) / v
    assert not np.isnan(
        test
    ), f"Unexpected test value, test={test}, signal={signal}, m={m}, v={v}"
    trend = np.sign(test) if np.abs(test) > threshold else 0
    return trend, test


def filter_data(
    df: pandas.DataFrame,
    filter_in: Optional[str] = None,
    filter_out: Optional[str] = None,
    verbose: int = 0,
) -> pandas.DataFrame:
    """
    Argument `filter` follows the syntax
    ``<column1>:<fmt1>//<column2>:<fmt2>``.

    The format is the following:

    * a value or a set of values separated by ``;``
    """
    if not filter_in and not filter_out:
        return df

    def _f(fmt):
        cond = {}
        if isinstance(fmt, str):
            cols = fmt.split("//")
            for c in cols:
                assert ":" in c, f"Unexpected value {c!r} in fmt={fmt!r}, cols={cols!r}"
                spl = c.split(":")
                assert (
                    len(spl) == 2
                ), f"Unexpected value {c!r} in fmt={fmt!r}, spl={spl}, cols={cols}"
                name, fil = spl
                cond[name] = set(fil.split(";"))
        return cond

    if filter_in:
        cond = _f(filter_in)
        assert isinstance(cond, dict), f"Unexpected type {type(cond)} for fmt={filter_in!r}"
        for k, v in cond.items():
            if k not in df.columns:
                continue
            if verbose:
                print(
                    f"[_filter_data] filter in column {k!r}, "
                    f"values {v!r} among {set(df[k].astype(str))}"
                )
            df = df[df[k].astype(str).isin(v)]

    if filter_out:
        cond = _f(filter_out)
        assert isinstance(cond, dict), f"Unexpected type {type(cond)} for fmt={filter_out!r}"
        for k, v in cond.items():
            if k not in df.columns:
                continue
            if verbose:
                print(
                    f"[_filter_data] filter out column {k!r}, "
                    f"values {v!r} among {set(df[k].astype(str))}"
                )
            df = df[~df[k].astype(str).isin(v)]
    return df


def enumerate_csv_files(
    data: Union[
        pandas.DataFrame, List[Union[str, Tuple[str, str]]], str, Tuple[str, str, str, str]
    ],
    verbose: int = 0,
    filtering: Optional[Callable[[str], bool]] = None,
) -> Iterator[Union[pandas.DataFrame, str, Tuple[str, str, str, str]]]:
    """
    Enumerates files considered for the aggregation.
    Only csv files are considered.
    If a zip file is given, the function digs into the zip files and
    loops over csv candidates.

    :param data: dataframe with the raw data or a file or list of files
    :param verbose: verbosity
    :param filtering: function to filter in or out files in zip files,
        must return true to keep the file, false to skip it.
    :return: a generator yielding tuples with the filename, date, full path and zip file

    data can contains:
    * a dataframe
    * a string for a filename, zip or csv
    * a list of string
    * a tuple
    """
    if not isinstance(data, list):
        data = [data]
    for itn, filename in enumerate(data):  # pyrefly: ignore[bad-argument-type]
        if isinstance(filename, pandas.DataFrame):
            if verbose:
                print(f"[enumerate_csv_files] data[{itn}] is a dataframe")
            yield filename
            continue

        if isinstance(filename, tuple):
            # A file in a zipfile
            if verbose:
                print(f"[enumerate_csv_files] data[{itn}] is {filename!r}")
            yield filename
            continue

        if os.path.exists(filename):
            ext = os.path.splitext(filename)[-1]
            if ext == ".csv":
                # We check the first line is ok.
                if verbose:
                    print(f"[enumerate_csv_files] data[{itn}] is a csv file: {filename!r}]")
                dt = datetime.datetime.fromtimestamp(os.stat(filename).st_mtime)
                du = dt.strftime("%Y-%m-%d %H:%M:%S")
                yield (os.path.split(filename)[-1], du, filename, "")
                continue

            if ext == ".zip":
                if verbose:
                    print(f"[enumerate_csv_files] data[{itn}] is a zip file: {filename!r}]")
                zf = zipfile.ZipFile(filename, "r")
                for ii, info in enumerate(zf.infolist()):
                    name = info.filename
                    if filtering is None:
                        ext = os.path.splitext(name)[-1]
                        if ext != ".csv":
                            continue
                    elif not filtering(name):
                        continue
                    if verbose:
                        print(f"[enumerate_csv_files] data[{itn}][{ii}] is a csv file: {name!r}]")
                    with zf.open(name) as zzf:
                        first_line = zzf.readline()
                    if b"," not in first_line:
                        continue
                    yield (
                        os.path.split(name)[-1],
                        "%04d-%02d-%02d %02d:%02d:%02d" % info.date_time,
                        name,
                        filename,
                    )
                zf.close()
                continue

            raise AssertionError(f"Unexpected format {filename!r}, cannot read it.")

        # filename is a pattern.
        found = glob.glob(filename)
        if verbose and not found:
            print(f"[enumerate_csv_files] unable to find file in {filename!r}")
        for ii, f in enumerate(found):
            if verbose:
                print(f"[enumerate_csv_files] data[{itn}][{ii}] {f!r} from {filename!r}")
            yield from enumerate_csv_files(f, verbose=verbose, filtering=filtering)


def open_dataframe(
    data: Union[str, Tuple[str, str, str, str], pandas.DataFrame],
) -> pandas.DataFrame:
    """
    Opens a filename defined by function
    :func:`yobx.helpers._log_helper.enumerate_csv_files`.

    :param data: a dataframe, a filename, a tuple indicating the file is coming
        from a zip file
    :return: a dataframe
    """
    if isinstance(data, pandas.DataFrame):
        return data
    if isinstance(data, str):
        df = pandas.read_csv(data, low_memory=False)
        df["RAWFILENAME"] = data
        return df
    if isinstance(data, tuple):
        if not data[-1]:
            df = pandas.read_csv(data[2], low_memory=False)
            df["RAWFILENAME"] = data[2]
            return df
        zf = zipfile.ZipFile(data[-1])
        with zf.open(data[2]) as f:
            df = pandas.read_csv(f, low_memory=False)
            df["RAWFILENAME"] = f"{data[-1]}/{data[2]}"
        zf.close()
        return df

    raise ValueError(f"Unexpected value for data: {data!r}")


def align_dataframe_with(
    df: pandas.DataFrame, baseline: pandas.DataFrame, fill_value: float = 0
) -> Optional[pandas.DataFrame]:
    """
    Modifies the first dataframe *df* to get the exact same number of columns and rows.
    They must share the same levels on both axes. Empty cells are filled with 0.
    We only keep the numerical columns. The function return None if the output is empty.
    """
    df = df.select_dtypes(include="number")
    if df.shape[1] == 0:
        return None
    bool_cols = list(df.select_dtypes(include="bool").columns)
    if bool_cols:
        df[bool_cols] = df[bool_cols].astype(int)
    assert df.columns.names == baseline.columns.names or df.index.names == baseline.index.names, (
        f"Levels mismatch, expected index.names={baseline.index.names}, "
        f"expected columns.names={baseline.columns.names}, "
        f"got index.names={df.index.names}, "
        f"got columns.names={df.columns.names}"
    )
    dtypes = set(df[c].dtype for c in df.columns)
    assert all(np.issubdtype(dt, np.number) for dt in dtypes), (
        f"All columns in the first dataframe are expected to share "
        f"the same type or be at least numerical but got {dtypes}\n{df}"
    )
    common_index = df.index.intersection(baseline.index)
    cp = pandas.DataFrame(float(fill_value), index=baseline.index, columns=baseline.columns)
    for c in df.columns:
        if c not in cp.columns or not np.issubdtype(df[c].dtype, np.number):
            continue
        cp.loc[common_index, c] = df.loc[common_index, c].astype(cp[c].dtype)
    return cp


def apply_excel_style(
    filename_or_writer: Any,
    f_highlights: Optional[  # type: ignore[name-defined]
        Dict[
            str,
            Callable[
                [Any], "CubeViewDef.HighLightKind"  # pyrefly: ignore[unknown-name]]  # noqa: F821
            ],
        ]
    ] = None,
    time_mask_view: Optional[Dict[str, pandas.DataFrame]] = None,
    verbose: int = 0,
):
    """
    Applies styles on all sheets in a file unless the sheet is too big.

    :param filename_or_writer: filename, modified inplace
    :param f_highlight: color function to apply, one per sheet
    :param time_mask_view: if specified, it contains dataframe with the same shape
        and values in {-1, 0, +1} which indicates if a value is unexpectedly lower (-1)
        or higher (+1), it changes the color of the background then.
    :param verbose: progress loop
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill
    from .cube_helper import CubeViewDef

    if isinstance(filename_or_writer, str):
        workbook = load_workbook(filename_or_writer)
        save = True
    else:
        workbook = filename_or_writer.book
        save = False

    mask_low = PatternFill(fgColor="AAAAF0", fill_type="solid")
    mask_high = PatternFill(fgColor="F0AAAA", fill_type="solid")

    left = Alignment(horizontal="left")
    left_shrink = Alignment(horizontal="left", shrink_to_fit=True)
    right = Alignment(horizontal="right")
    font_colors = {
        CubeViewDef.HighLightKind.GREEN: Font(color="00AA00"),
        CubeViewDef.HighLightKind.RED: Font(color="FF0000"),
    }
    if verbose:  # pragma: no cover
        from tqdm import tqdm

        sheet_names = tqdm(list(workbook.sheetnames))
    else:
        sheet_names = workbook.sheetnames
    for name in sheet_names:
        if time_mask_view and name in time_mask_view:
            mask = time_mask_view[name]
            with pandas.ExcelWriter(io.BytesIO(), engine="openpyxl") as mask_writer:
                mask.to_excel(mask_writer, sheet_name=name)
                sheet_mask = mask_writer.sheets[name]
        else:
            sheet_mask = None

        f_highlight = f_highlights.get(name, None) if f_highlights else None
        sheet = workbook[name]
        n_rows = sheet.max_row
        n_cols = sheet.max_column
        if n_rows * n_cols > 2**16 or n_rows > 2**13:
            # Too big.
            continue
        co: Dict[int, int] = {}
        sizes: Dict[int, int] = {}
        cols = set()
        for i in range(1, n_rows + 1):
            for j, cell in enumerate(sheet[i]):
                if j > n_cols:
                    break
                cols.add(cell.column)
                if isinstance(cell.value, float):
                    co[j] = co.get(j, 0) + 1
                elif isinstance(cell.value, str):
                    sizes[cell.column] = max(sizes.get(cell.column, 0), len(cell.value))

        for k, v in sizes.items():
            c = get_column_letter(k)
            sheet.column_dimensions[c].width = min(max(8, v), 30)
        for k in cols:
            if k not in sizes:
                c = get_column_letter(k)
                sheet.column_dimensions[c].width = 15

        for i in range(1, n_rows + 1):
            for j, cell in enumerate(sheet[i]):
                if j > n_cols:
                    break
                if isinstance(cell.value, pandas.Timestamp):
                    cell.alignment = right
                    dt = cell.value.to_pydatetime()
                    cell.value = dt
                    cell.number_format = (
                        "YYYY-MM-DD"
                        if (
                            dt.hour == 0
                            and dt.minute == 0
                            and dt.second == 0
                            and dt.microsecond == 0
                        )
                        else "YYYY-MM-DD 00:00:00"
                    )
                elif isinstance(cell.value, (float, int)):
                    cell.alignment = right
                    x = abs(cell.value)
                    if int(x) == x:
                        cell.number_format = "0"
                    elif x > 5000:
                        cell.number_format = "# ##0"
                    elif x >= 500:
                        cell.number_format = "0.0"
                    elif x >= 50:
                        cell.number_format = "0.00"
                    elif x >= 5:
                        cell.number_format = "0.000"
                    elif x > 0.5:
                        cell.number_format = "0.0000"
                    elif x > 0.005:
                        cell.number_format = "0.00000"
                    else:
                        cell.number_format = "0.000E+00"
                    if f_highlight:
                        h = f_highlight(cell.value)
                        if h in font_colors:
                            cell.font = font_colors[h]
                elif isinstance(cell.value, str) and len(cell.value) > 70:
                    cell.alignment = left_shrink
                else:
                    cell.alignment = left
                    if f_highlight:
                        h = f_highlight(cell.value)
                        if h in font_colors:
                            cell.font = font_colors[h]

        if sheet_mask is not None:
            for i in range(1, n_rows + 1):
                for j, (cell, cell_mask) in enumerate(zip(sheet[i], sheet_mask[i])):
                    if j > n_cols:
                        break
                    if cell_mask.value not in (1, -1):
                        continue
                    cell.fill = mask_low if cell_mask.value < 0 else mask_high

    if save:
        workbook.save(filename_or_writer)
